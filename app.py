import os
import re
from flask import Flask, render_template, request, redirect, url_for, send_from_directory
from docx import Document
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# إعداد المسارات
UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads')
EXCEL_FILE_PATH = os.path.join(UPLOAD_FOLDER, 'client_data.xlsx')

# التأكد من أن المجلد 'uploads' موجود
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# التأكد من أن ملف Excel موجود، وإذا لم يكن موجودًا، يتم إنشاؤه مع عناوين الأعمدة
if not os.path.exists(EXCEL_FILE_PATH):
    wb = Workbook()
    ws = wb.active
    ws.append(["الاسم", "البريد الإلكتروني", "رقم الهاتف"])  # إضافة العناوين
    wb.save(EXCEL_FILE_PATH)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # الحصول على البيانات من النموذج
        name = request.form['name']
        email = request.form['email']
        phone = request.form['phone']

        # تحديد المسار المطلق للقالب
        doc_path = os.path.join(os.getcwd(), 'templates', 'template_document.docx')

        # التحقق من وجود القالب قبل فتحه
        if not os.path.exists(doc_path):
            return "⚠️ الملف template_document.docx غير موجود في مجلد templates."

        # فتح المستند واستبدال النصوص
        doc = Document(doc_path)
        for paragraph in doc.paragraphs:
            paragraph.text = re.sub(r'اسم العميل', name, paragraph.text)
            paragraph.text = re.sub(r'البريد الإلكتروني', email, paragraph.text)
            paragraph.text = re.sub(r'رقم الهاتف', phone, paragraph.text)

        # تحديد المسار لحفظ المستند المعدل
        updated_doc_path = os.path.join(UPLOAD_FOLDER, 'updated_document.docx')

        # حفظ المستند بعد التعديل
        doc.save(updated_doc_path)

        # حفظ البيانات في ملف Excel
        try:
            wb = load_workbook(EXCEL_FILE_PATH)
            ws = wb.active
            ws.append([name, email, phone])  # إضافة البيانات الجديدة
            wb.save(EXCEL_FILE_PATH)
        except Exception as e:
            return f"⚠️ حدث خطأ أثناء تحديث ملف Excel: {e}"

        # إعادة توجيه المستخدم إلى صفحة نجاح التحميل
        return redirect(url_for('download_success'))

    return render_template('index.html')

@app.route('/download_success')
def download_success():
    """ صفحة نجاح بعد تحميل الملف """
    return render_template('download_success.html')

@app.route('/download/<filename>')
def download(filename):
    """ إرسال الملف المعدل للمستخدم """
    return send_from_directory(UPLOAD_FOLDER, filename)

if __name__ == '__main__':
    app.run(debug=True)
